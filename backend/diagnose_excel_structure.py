"""
엑셀 파일의 실제 구조를 진단하는 스크립트
병합된 셀, 다중 행 헤더 등을 확인
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

def diagnose_excel_structure():
    print("=" * 80)
    print("🔍 HPLM 엑셀 파일 구조 진단")
    print("=" * 80)
    
    excel_path = "uploads/HPLM 솔루션 과제정보.xlsx"
    
    print(f"\n📄 파일: {excel_path}\n")
    
    # 1. openpyxl로 원본 구조 확인
    print("1️⃣ 원본 셀 구조 확인 (openpyxl)")
    print("-" * 80)
    
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        sheet_name = ws.title
        
        print(f"시트명: {sheet_name}")
        print(f"최대 행: {ws.max_row}")
        print(f"최대 열: {ws.max_column}")
        
        # 처음 5행의 실제 셀 값 출력
        print(f"\n처음 5행의 원본 데이터:")
        print("-" * 80)
        
        for row_idx in range(1, min(6, ws.max_row + 1)):
            print(f"\n행 {row_idx}:")
            row_data = []
            for col_idx in range(1, min(20, ws.max_column + 1)):  # 처음 20열만
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value:
                    row_data.append(f"  열{col_idx}: {value}")
            
            if row_data:
                print("\n".join(row_data))
            else:
                print("  (빈 행)")
        
        # 병합된 셀 확인
        print(f"\n\n병합된 셀 정보:")
        print("-" * 80)
        if ws.merged_cells:
            for merged_range in list(ws.merged_cells.ranges)[:10]:  # 처음 10개만
                print(f"  {merged_range}")
        else:
            print("  병합된 셀 없음")
        
    except Exception as e:
        print(f"❌ openpyxl 오류: {e}")
    
    # 2. pandas로 읽은 결과 확인
    print(f"\n\n2️⃣ pandas 기본 읽기 결과")
    print("-" * 80)
    
    try:
        df = pd.read_excel(excel_path, sheet_name=0)
        print(f"행 개수: {len(df)}")
        print(f"열 개수: {len(df.columns)}")
        print(f"\n컬럼명:")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i}. {col}")
        
        print(f"\n처음 3행 데이터:")
        print(df.head(3).to_string())
        
    except Exception as e:
        print(f"❌ pandas 오류: {e}")
    
    # 3. 헤더 행 감지 시도
    print(f"\n\n3️⃣ 다양한 헤더 행 옵션 테스트")
    print("-" * 80)
    
    for header_row in [0, 1, [0, 1], None]:
        try:
            print(f"\nheader={header_row} 옵션:")
            df = pd.read_excel(excel_path, sheet_name=0, header=header_row)
            print(f"  행 개수: {len(df)}")
            print(f"  열 개수: {len(df.columns)}")
            print(f"  컬럼명 (처음 5개): {list(df.columns)[:5]}")
            
            # "물류형 BCR" 찾기
            if '과제명' in df.columns:
                bcr_rows = df[df['과제명'].astype(str).str.contains('BCR', na=False)]
                if not bcr_rows.empty:
                    print(f"  ✅ '물류형 BCR' 발견:")
                    for col in df.columns:
                        val = bcr_rows.iloc[0][col]
                        if pd.notna(val):
                            print(f"    {col}: {val}")
        except Exception as e:
            print(f"  ❌ 오류: {e}")
    
    # 4. 물류형 BCR 행 상세 분석
    print(f"\n\n4️⃣ '물류형 BCR' 행 상세 분석")
    print("-" * 80)
    
    try:
        # 기본 읽기
        df = pd.read_excel(excel_path, sheet_name=0)
        
        # "물류" 또는 "BCR" 포함된 행 찾기
        for idx, row in df.iterrows():
            row_str = ' '.join([str(v) for v in row.values if pd.notna(v)])
            if 'BCR' in row_str or '물류' in row_str:
                print(f"\n행 {idx + 2} (Excel 기준) 발견:")
                print("-" * 40)
                for col_name, value in row.items():
                    if pd.notna(value):
                        print(f"  {col_name}: {value}")
                print()
    
    except Exception as e:
        print(f"❌ 오류: {e}")

if __name__ == "__main__":
    diagnose_excel_structure()

